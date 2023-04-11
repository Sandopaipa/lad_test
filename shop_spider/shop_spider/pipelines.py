# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
from itemadapter import ItemAdapter
import openpyxl


class ShopSpiderPipeline:
    def __init__(self):
        self.book = openpyxl.Workbook()
        self.sheet = self.book.active
        self.sheet['A1'] = 'Name'
        self.sheet['B1'] = 'Price'

    def process_item(self, item, spider):
        row_num = self.sheet.max_row + 1
        self.sheet.cell(row=row_num, column=1).value = item['name']
        self.sheet.cell(row=row_num, column=2).value = item['price']
        self.book.save('output.xls')
        return item
